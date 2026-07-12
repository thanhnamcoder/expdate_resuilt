"""
Module dùng chung để xử lý dữ liệu của 1 WriteOffBatch: tạo file Excel (danh
sách items + giá cost + tổng) và đọc danh sách ảnh đính kèm từ ổ đĩa.

Tái sử dụng ở cả:
- mail_api.SendEmailAPIView (đính kèm rời từng file vào email)
- item_views.WriteOffBatchExportView (đóng gói zip để tải về máy)

Tách riêng để không phải viết lại cùng 1 logic build Excel/đọc ảnh ở 2 nơi.
"""
import io
import os
import re
import mimetypes

from openpyxl import Workbook
from django.conf import settings

from .models import WriteOffItem


def get_safe_batch_name(batch):
    """Tên batch đã loại bỏ ký tự không hợp lệ để dùng làm tên file."""
    return re.sub(r'[\\/:*?"<>|]', '_', batch.name or 'writeoff_batch')


def build_writeoff_excel_bytes(batch):
    """Tạo file Excel (bytes) chứa danh sách items của batch, có cột
    Giá cost / Tổng định dạng number, kèm dòng tổng cộng cuối cùng."""
    items = WriteOffItem.objects.filter(writeoff_batch=batch).order_by('id')

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Items'
    sheet.append(['Barcode', 'Tên hàng', 'Item Code', 'Số lượng', 'Giá cost', 'Tổng'])

    for item in items:
        total = (item.unit_cost or 0) * item.quantity if item.unit_cost else 0
        row_idx = sheet.max_row + 1
        sheet.append([
            item.barcode,
            item.itemname,
            item.item_code or '',
            item.quantity,
            float(item.unit_cost) if item.unit_cost else 0,
            float(total)
        ])
        sheet.cell(row=row_idx, column=5).number_format = '#,##0'
        sheet.cell(row=row_idx, column=6).number_format = '#,##0'

    if items.exists():
        sheet.append([])
        total_row_idx = sheet.max_row + 1
        sheet.append(['', '', '', 'Tổng:', '', float(batch.total_cost) if batch.total_cost is not None else 0])
        sheet.cell(row=total_row_idx, column=6).number_format = '#,##0'

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()


def get_writeoff_batch_images(batch):
    """Đọc từng file ảnh của batch trực tiếp từ ổ đĩa, trả về danh sách dict:
    {filename, maintype, subtype, data (bytes)} — sẵn sàng để đính kèm email
    hoặc ghi vào zip, không cần tải qua URL public."""
    images = []
    if not batch.file_paths:
        return images

    for file_path in batch.file_paths.split('|'):
        file_path = file_path.strip()
        if not file_path:
            continue

        full_path = os.path.join(settings.BASE_DIR, file_path.lstrip('/'))
        if not os.path.exists(full_path):
            continue

        mime_type, _ = mimetypes.guess_type(full_path)
        if mime_type:
            maintype, subtype = mime_type.split('/', 1)
        else:
            maintype, subtype = 'application', 'octet-stream'

        with open(full_path, 'rb') as f:
            data = f.read()

        images.append({
            'filename': os.path.basename(full_path),
            'maintype': maintype,
            'subtype': subtype,
            'data': data,
        })

    return images


def build_writeoff_export_data(batch):
    """Gộp cả 2 phần trên: trả về dict sẵn sàng dùng cho cả gửi email lẫn export zip.
    {
        'safe_name': str,
        'excel_bytes': bytes,
        'images': [{'filename', 'maintype', 'subtype', 'data'}, ...],
    }
    """
    return {
        'safe_name': get_safe_batch_name(batch),
        'excel_bytes': build_writeoff_excel_bytes(batch),
        'images': get_writeoff_batch_images(batch),
    }