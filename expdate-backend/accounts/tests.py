import json
from datetime import date, timedelta
from decimal import Decimal
from unittest.mock import patch

from django.contrib import admin
from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connections
from django.test import TestCase
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.test import APIRequestFactory, force_authenticate

from .admin import ItemAdmin
from .item_views import ItemBatchCreateView, ItemCreateView, ProductCostLookupView, ProductCostListView, WriteOffBatchDeleteView, WriteOffItemDeleteView
from .mail_api import SendEmailAPIView
from .models import Item, PackingList, ProductCost, ProductData, WriteOffArchive, WriteOffBatch, WriteOffItem


class ItemAdminImportCostTests(TestCase):
    def test_item_admin_registers_import_cost_url(self):
        item_admin = ItemAdmin(model=ProductData, admin_site=admin.site)

        urls = item_admin.get_urls()

        self.assertTrue(any(getattr(url, 'name', None) == 'import_cost' for url in urls))

    def test_prepare_cost_import_workbook_keeps_only_target_sheets_and_removes_columns_and_rows(self):
        item_admin = ItemAdmin(model=ProductData, admin_site=admin.site)
        workbook = Workbook()
        target_sheet = workbook.active
        target_sheet.title = 'Physical Inventory Result (P2)'
        for row_idx in range(1, 13):
            for col_idx in range(1, 21):
                target_sheet.cell(row=row_idx, column=col_idx, value=f'{row_idx}-{col_idx}')

        workbook.create_sheet('Physical Inventory Result (P3)')
        workbook.create_sheet('Other Sheet')

        processed_workbook = item_admin._prepare_cost_import_workbook(workbook)

        self.assertEqual(processed_workbook.sheetnames, ['Physical Inventory Result (P2)'])
        self.assertEqual(processed_workbook['Physical Inventory Result (P2)'].max_column, 8)
        self.assertEqual(processed_workbook['Physical Inventory Result (P2)'].max_row, 1)

    def test_prepare_cost_import_workbook_keeps_the_first_column(self):
        item_admin = ItemAdmin(model=ProductData, admin_site=admin.site)
        workbook = Workbook()
        target_sheet = workbook.active
        target_sheet.title = 'Physical Inventory Result (P2)'
        for row_idx in range(1, 13):
            for col_idx in range(1, 21):
                target_sheet.cell(row=row_idx, column=col_idx, value=f'{row_idx}-{col_idx}')

        processed_workbook = item_admin._prepare_cost_import_workbook(workbook)
        processed_sheet = processed_workbook['Physical Inventory Result (P2)']

        self.assertEqual(processed_sheet.cell(row=1, column=1).value, '9-1')


class ProductCostLookupTests(TestCase):
    def test_lookup_returns_cost_for_existing_item_code(self):
        ProductCost.objects.create(item_code='ITEM-001', item_name='Test Item', unit_cost='12.50')

        view = ProductCostLookupView.as_view()
        request = APIRequestFactory().get('/api/accounts/product-cost/ITEM-001/')

        response = view(request, item_code='ITEM-001')

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['item_code'], 'ITEM-001')
        self.assertEqual(response.data['item_name'], 'Test Item')
        self.assertEqual(response.data['unit_cost'], 12.5)

    def test_list_returns_all_product_costs(self):
        ProductCost.objects.create(item_code='ITEM-001', item_name='Test Item 1', unit_cost='12.50')
        ProductCost.objects.create(item_code='ITEM-002', item_name='Test Item 2', unit_cost='8.00')

        view = ProductCostListView.as_view()
        request = APIRequestFactory().get('/api/accounts/product-cost/')

        response = view(request)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['count'], 2)
        self.assertEqual(response.data['data'][0]['item_code'], 'ITEM-001')
        self.assertEqual(response.data['data'][1]['item_code'], 'ITEM-002')


class ItemCreateViewPackingListTests(TestCase):
    databases = {'default', 'sqlite'}

    def setUp(self):
        self.user = User.objects.create_user(username='tester', password='secret123')
        self.factory = APIRequestFactory()
        with connections['sqlite'].cursor() as cursor:
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS PackingList (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    item_code VARCHAR(100),
                    barcode VARCHAR(100) NOT NULL,
                    description VARCHAR(255) NOT NULL,
                    exp_date VARCHAR(20) NOT NULL,
                    quantity INTEGER NOT NULL
                )
                """
            )

    def test_create_item_updates_existing_packing_list_when_quantity_is_string(self):
        PackingList.objects.using('sqlite').create(
            item_code='',
            barcode='barcode-1',
            description='old item',
            exp_date='03/07/2026',
            quantity='7',
        )

        view = ItemCreateView.as_view()
        request = self.factory.post(
            '/api/items/',
            {
                'barcode': 'barcode-1',
                'itemname': 'New item',
                'quantity': '3',
                'expdate': '03/07/2026',
                'username': 'tester',
                'item_code': '',
                'stocktake': False,
            },
            format='json',
        )
        force_authenticate(request, user=self.user)

        response = view(request)

        self.assertEqual(response.status_code, 201)
        packing = PackingList.objects.using('sqlite').filter(
            item_code='',
            barcode='barcode-1',
            exp_date='03/07/2026',
        ).first()
        self.assertIsNotNone(packing)
        self.assertEqual(packing.quantity, 10)
        self.assertEqual(packing.description, 'New item')


class ItemCreateViewWriteoffTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='tester', password='secret123')
        self.factory = APIRequestFactory()

    def test_create_writeoff_item_saves_upload_and_record(self):
        uploaded_file = SimpleUploadedFile('photo.jpg', b'fake-image-bytes', content_type='image/jpeg')
        view = ItemCreateView.as_view()
        request = self.factory.post(
            '/api/items/',
            {
                'barcode': 'wo-1',
                'itemname': 'Write off item',
                'quantity': '2',
                'username': 'tester',
                'item_code': '',
                'writeoff': 'true',
                'stocktake': 'false',
                'files': [uploaded_file],
            },
            format='multipart',
        )
        force_authenticate(request, user=self.user)

        response = view(request)

        self.assertEqual(response.status_code, 201)
        writeoff_item = WriteOffItem.objects.filter(barcode='wo-1').first()
        self.assertIsNotNone(writeoff_item)
        self.assertEqual(writeoff_item.quantity, 2)
        self.assertIsNotNone(writeoff_item.writeoff_batch)
        self.assertEqual(writeoff_item.writeoff_batch.file_paths.count('Image/'), 1)

    def test_get_items_by_barcode_returns_search_results(self):
        Item.objects.create(
            barcode='43',
            itemname='Test Item',
            quantity=5,
            expdate='2026-07-08',
            user=self.user,
            item_code='ITEM-43',
            stocktake=False,
            writeoff=False,
        )

        view = ItemCreateView.as_view()
        request = self.factory.get('/api/items/?barcode=43&username=tester&date=2026-07-08')
        force_authenticate(request, user=self.user)

        response = view(request)

        self.assertEqual(response.status_code, 200)
        self.assertIn('data', response.data)
        self.assertEqual(len(response.data['data']), 1)
        self.assertEqual(response.data['data'][0]['barcode'], '43')
        self.assertEqual(response.data['data'][0]['itemname'], 'Test Item')
        self.assertEqual(response.data['data'][0]['quantity'], 5)
        self.assertEqual(response.data['data'][0]['item_code'], 'ITEM-43')

    def test_create_writeoff_item_saves_total_cost_to_batch(self):
        uploaded_file = SimpleUploadedFile('photo.jpg', b'fake-image-bytes', content_type='image/jpeg')
        view = ItemCreateView.as_view()
        request = self.factory.post(
            '/api/items/',
            {
                'barcode': 'wo-3',
                'itemname': 'Write off item with cost',
                'quantity': '2',
                'username': 'tester',
                'item_code': '',
                'writeoff': 'true',
                'stocktake': 'false',
                'unit_cost': '5.50',
                'files': [uploaded_file],
            },
            format='multipart',
        )
        force_authenticate(request, user=self.user)

        response = view(request)

        self.assertEqual(response.status_code, 201)
        batch = WriteOffBatch.objects.get(name='Write off item with cost')
        self.assertEqual(batch.total_cost, Decimal('11.00'))

    def test_create_writeoff_item_without_file_is_rejected(self):
        view = ItemCreateView.as_view()
        request = self.factory.post(
            '/api/items/',
            {
                'barcode': 'wo-2',
                'itemname': 'Write off item no file',
                'quantity': '1',
                'username': 'tester',
                'item_code': '',
                'writeoff': 'true',
                'stocktake': 'false',
            },
            format='multipart',
        )
        force_authenticate(request, user=self.user)

        response = view(request)

        self.assertEqual(response.status_code, 400)
        self.assertFalse(WriteOffItem.objects.filter(barcode='wo-2').exists())


class WriteOffBatchCostTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='tester', password='secret123')
        self.factory = APIRequestFactory()

    def test_batch_create_saves_total_cost_to_writeoff_batch(self):
        uploaded_file = SimpleUploadedFile('photo.jpg', b'fake-image-bytes', content_type='image/jpeg')
        view = ItemBatchCreateView.as_view()
        request = self.factory.post(
            '/api/accounts/items/writeoff_batches/',
            {
                'payload': json.dumps({
                    'WO Batch': [
                        {
                            'barcode': 'wo-cost-1',
                            'itemname': 'Costed Item',
                            'quantity': 2,
                            'item_code': 'ITEM-001',
                            'unit_cost': '5.50',
                        }
                    ]
                }),
                'files': [uploaded_file],
            },
            format='multipart',
        )
        force_authenticate(request, user=self.user)

        response = view(request)

        self.assertEqual(response.status_code, 201)
        batch = WriteOffBatch.objects.get(name='WO Batch')
        self.assertEqual(batch.total_cost, Decimal('11.00'))

    def test_batch_create_saves_unit_cost_to_writeoff_item(self):
        uploaded_file = SimpleUploadedFile('photo.jpg', b'fake-image-bytes', content_type='image/jpeg')
        view = ItemBatchCreateView.as_view()
        request = self.factory.post(
            '/api/accounts/items/writeoff_batches/',
            {
                'payload': json.dumps({
                    'WO Batch': [
                        {
                            'barcode': 'wo-unit-cost-1',
                            'itemname': 'Costed Item',
                            'quantity': 1,
                            'item_code': 'ITEM-002',
                            'unit_cost': '6686.35',
                        }
                    ]
                }),
                'files': [uploaded_file],
            },
            format='multipart',
        )
        force_authenticate(request, user=self.user)

        response = view(request)

        self.assertEqual(response.status_code, 201)
        writeoff_item = WriteOffItem.objects.get(barcode='wo-unit-cost-1')
        self.assertEqual(writeoff_item.unit_cost, Decimal('6686.35'))


class WriteOffDeletionTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username='tester', password='secret123')
        self.factory = APIRequestFactory()

    @patch('accounts.mail_api.smtplib.SMTP')
    def test_send_email_archives_batch_after_success(self, mock_smtp):
        batch = WriteOffBatch.objects.create(user=self.user, name='batch-email-archive')
        item = WriteOffItem.objects.create(
            writeoff_batch=batch,
            barcode='wo-email-archive',
            itemname='Email Archive Item',
            quantity=2,
        )

        view = SendEmailAPIView.as_view()
        request = self.factory.post(
            '/api/accounts/send-email/',
            {
                'to_email': 'tester@example.com',
                'subject': 'Test',
                'body_html': '<p>Hi</p>',
                'batch_id': batch.id,
            },
            format='json',
        )
        force_authenticate(request, user=self.user)

        response = view(request)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(WriteOffBatch.objects.filter(id=batch.id).exists())
        self.assertFalse(WriteOffItem.objects.filter(id=item.id).exists())
        self.assertTrue(WriteOffArchive.objects.filter(record_type='batch', source_id=batch.id).exists())
        self.assertTrue(WriteOffArchive.objects.filter(record_type='item', source_id=item.id).exists())
        mock_smtp.assert_called_once()

    def test_delete_writeoff_item_archives_item_and_keeps_batch(self):
        batch = WriteOffBatch.objects.create(user=self.user, name='batch-delete-item')
        item = WriteOffItem.objects.create(
            writeoff_batch=batch,
            barcode='wo-delete-item',
            itemname='Item to delete',
            quantity=2,
        )

        view = WriteOffItemDeleteView.as_view()
        request = self.factory.delete(f'/api/items/writeoff_items/{item.id}/delete/', format='json')
        force_authenticate(request, user=self.user)

        response = view(request, item_id=item.id)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(WriteOffItem.objects.filter(id=item.id).exists())
        self.assertTrue(WriteOffBatch.objects.filter(id=batch.id).exists())
        self.assertTrue(WriteOffArchive.objects.filter(record_type='item', source_id=item.id).exists())

    def test_delete_writeoff_batch_archives_batch_and_items(self):
        batch = WriteOffBatch.objects.create(user=self.user, name='batch-delete-all')
        item_1 = WriteOffItem.objects.create(
            writeoff_batch=batch,
            barcode='wo-delete-1',
            itemname='Item 1',
            quantity=1,
        )
        item_2 = WriteOffItem.objects.create(
            writeoff_batch=batch,
            barcode='wo-delete-2',
            itemname='Item 2',
            quantity=3,
        )

        view = WriteOffBatchDeleteView.as_view()
        request = self.factory.delete(f'/api/items/writeoff_batches/{batch.id}/delete/', format='json')
        force_authenticate(request, user=self.user)

        response = view(request, batch_id=batch.id)

        self.assertEqual(response.status_code, 200)
        self.assertFalse(WriteOffBatch.objects.filter(id=batch.id).exists())
        self.assertFalse(WriteOffItem.objects.filter(writeoff_batch=batch).exists())
        self.assertTrue(WriteOffArchive.objects.filter(record_type='batch', source_id=batch.id).exists())
        self.assertTrue(WriteOffArchive.objects.filter(record_type='item', source_id=item_1.id).exists())
        self.assertTrue(WriteOffArchive.objects.filter(record_type='item', source_id=item_2.id).exists())

    def test_cleanup_old_archives_removes_entries_older_than_one_month(self):
        old_entry = WriteOffArchive.objects.create(
            record_type='batch',
            source_id=999,
            user=self.user,
            name='old batch',
        )
        recent_entry = WriteOffArchive.objects.create(
            record_type='batch',
            source_id=1000,
            user=self.user,
            name='recent batch',
        )
        WriteOffArchive.objects.filter(id=old_entry.id).update(deleted_at=timezone.now() - timedelta(days=40))
        WriteOffArchive.objects.filter(id=recent_entry.id).update(deleted_at=timezone.now() - timedelta(days=10))

        WriteOffArchive.cleanup_old_archives()

        self.assertFalse(WriteOffArchive.objects.filter(id=old_entry.id).exists())
        self.assertTrue(WriteOffArchive.objects.filter(id=recent_entry.id).exists())
