from django import forms
from django.contrib import admin
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin
from django.utils.translation import gettext_lazy as _
from django.urls import path
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.admin import site
from django.contrib import messages
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
from datetime import datetime
from .models import Profile, Item, ProductData
from django.db import transaction
from decimal import Decimal, InvalidOperation
# Monkey patch: Thay đổi hành vi của 'View site' trên admin
site.site_url = 'https://square-shark-greatly.ngrok-free.app/'

class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(
        label='Chọn file Excel',
        help_text='Chỉ chấp nhận file .xlsx'
    )

class ProfileInlineForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Lấy danh sách group duy nhất từ các profile đã có
        group_choices = [(g, g) for g in Profile.objects.values_list('group', flat=True).distinct() if g]
        group_choices.insert(0, ('', 'None'))  # Thêm lựa chọn None
        self.fields['group'].widget = forms.Select(choices=group_choices)

    class Meta:
        model = Profile
        fields = ('fullname', 'group')

class ProfileInline(admin.StackedInline):
    model = Profile
    can_delete = False
    verbose_name_plural = 'Profile'
    fields = ('fullname', 'group')
    form = ProfileInlineForm

class ItemInline(admin.TabularInline):
    model = Item
    extra = 0
    fields = ('barcode', 'itemname', 'quantity', 'expdate')
    readonly_fields = ('barcode', 'itemname', 'quantity', 'expdate')
    can_delete = False
    show_change_link = True
    
    def has_add_permission(self, request, obj=None):
        return False

class CustomUserAdmin(UserAdmin):
    inlines = (ProfileInline, ItemInline)
    list_display = ('username', 'email', 'user', 'manage', 'super', 'get_group', 'get_items_count')
    search_fields = ('username', 'email', 'profile__group', 'items__itemname', 'items__barcode')
    ordering = ('username',)

    def user(self, obj):
        return obj.is_active and not obj.is_staff and not obj.is_superuser
    user.boolean = True
    user.short_description = 'User'

    def manage(self, obj):
        return obj.is_staff and not obj.is_superuser
    manage.boolean = True
    manage.short_description = 'Manage'

    def super(self, obj):
        return obj.is_superuser
    super.boolean = True
    super.short_description = 'Super'

    def get_group(self, obj):
        return obj.profile.group if hasattr(obj, 'profile') else None
    get_group.short_description = 'Group'
    
    def get_items_count(self, obj):
        return obj.items.count()
    get_items_count.short_description = 'Items Count'

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)

    def save_formset(self, request, form, formset, change):
        instances = formset.save(commit=False)
        for instance in instances:
            if hasattr(instance, 'group'):
                Profile.objects.filter(user=instance.user).exclude(pk=instance.pk).delete()
            instance.save()
        formset.save_m2m()

class ItemAdmin(admin.ModelAdmin):
    list_display = ('barcode', 'itemname', 'quantity', 'expdate', 'get_user')
    search_fields = ('barcode', 'itemname', 'user__username')
    list_filter = ('expdate', 'user__profile__group')
    ordering = ('expdate',)
    change_list_template = 'admin/item_changelist.html'
    
    def get_user(self, obj):
        return obj.user.username if obj.user else None
    get_user.short_description = 'User'
    get_user.admin_order_field = 'user__username'

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel, name='import_excel'),
            path('import-cost/', self.import_cost, name='import_cost'),
        ]
        return my_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                if not excel_file.name.endswith(('.xlsx', '.xls')):
                    messages.error(request, 'File không phải là Excel (.xlsx)')
                    return HttpResponseRedirect("../")

                try:
                    wb = load_workbook(excel_file)
                    ws = wb.active
                    
                    # 1. Định nghĩa các cột bắt buộc
                    required_columns = {
                        'Item Barcode': None,
                        'Item Code': None,
                        'Item Name': None,
                        'Department': None,
                        'Category': None,
                        'Sub Category': None,
                        'Vendor Code': None,
                        'Vendor Name': None
                    }
                    
                    # 2. Tìm vị trí header (có thể nằm từ dòng 1-5)
                    header_row_index = None
                    header_values = None
                    
                    # Lấy 5 dòng đầu tiên
                    first_five_rows = list(ws.iter_rows(min_row=1, max_row=5))
                    
                    # Kiểm tra từng dòng để tìm header
                    for row_idx, row in enumerate(first_five_rows, start=1):
                        # Lấy giá trị của dòng hiện tại
                        current_row_values = [str(cell.value).strip() if cell.value else '' for cell in row]
                        
                        # Đếm số cột khớp với header cần thiết
                        matches = sum(1 for value in current_row_values if value in required_columns)
                        
                        # Nếu tìm thấy ít nhất 6 cột khớp (75% của 8 cột), coi đây là header
                        if matches >= 6:
                            header_row_index = row_idx
                            header_values = current_row_values
                            break
                    
                    if header_row_index is None:
                        messages.error(request, 'Không tìm thấy header trong 5 dòng đầu tiên của file')
                        return HttpResponseRedirect("../")
                    
                    # 3. Tìm vị trí của từng cột trong header
                    for idx, header_value in enumerate(header_values):
                        header_value = str(header_value).strip()
                        if header_value in required_columns:
                            required_columns[header_value] = idx
                    
                    # Kiểm tra các cột bắt buộc
                    missing_columns = [col for col, idx in required_columns.items() if idx is None]
                    if missing_columns:
                        messages.error(request, f'File thiếu các cột bắt buộc: {", ".join(missing_columns)}')
                        return HttpResponseRedirect("../")

                    # 4. Đọc và xử lý dữ liệu
                    success_count = 0
                    error_count = 0
                    error_rows = []
                    
                    # Lấy tất cả các dòng sau header
                    data_rows = list(ws.iter_rows(min_row=header_row_index + 1))
                    
                    with transaction.atomic():
                        products_to_create = []
                        existing_barcodes = set(ProductData.objects.values_list('item_barcode', flat=True))
                        duplicate_count = 0
                        for row_idx, row in enumerate(data_rows, start=header_row_index + 1):
                            try:
                                if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                                    continue
                                barcode = self._clean_cell_value(row[required_columns['Item Barcode']])
                                item_name = self._clean_cell_value(row[required_columns['Item Name']])
                                # Chỉ import nếu có đủ barcode và item_name
                                if not barcode or not item_name:
                                    error_count += 1
                                    error_rows.append(f'Dòng {row_idx}: Thiếu Barcode hoặc Item Name')
                                    continue
                                # Nếu barcode đã tồn tại thì bỏ qua và đếm
                                if barcode in existing_barcodes:
                                    duplicate_count += 1
                                    continue
                                product_data = ProductData(
                                    item_barcode=barcode,
                                    item_code=self._clean_cell_value(row[required_columns['Item Code']]) if required_columns['Item Code'] is not None else '',
                                    item_name=item_name,
                                    department=self._clean_cell_value(row[required_columns['Department']]) if required_columns['Department'] is not None else '',
                                    category=self._clean_cell_value(row[required_columns['Category']]) if required_columns['Category'] is not None else '',
                                    sub_category=self._clean_cell_value(row[required_columns['Sub Category']]) if required_columns['Sub Category'] is not None else '',
                                    vendor_code=self._clean_cell_value(row[required_columns['Vendor Code']]) if required_columns['Vendor Code'] is not None else '',
                                    vendor_name=self._clean_cell_value(row[required_columns['Vendor Name']]) if required_columns['Vendor Name'] is not None else ''
                                )
                                products_to_create.append(product_data)
                                existing_barcodes.add(barcode)
                                success_count += 1
                            except Exception as e:
                                error_count += 1
                                error_rows.append(f'Dòng {row_idx}: {str(e)}')
                                continue
                        if products_to_create:
                            ProductData.objects.bulk_create(products_to_create)
                        # Thông báo số lượng item bị bỏ qua do trùng lặp
                        if duplicate_count > 0:
                            messages.warning(request, f'Có {duplicate_count} item bị bỏ qua do trùng barcode đã tồn tại.')

                    # 5. Hiển thị kết quả
                    if success_count > 0:
                        messages.success(request, f'Import thành công {success_count} items')
                    if error_count > 0:
                        messages.warning(request, f'Có {error_count} dòng bị lỗi')
                        for error_row in error_rows:
                            messages.error(request, error_row)

                except Exception as e:
                    messages.error(request, f'Lỗi khi đọc file Excel: {str(e)}')
                
                return HttpResponseRedirect("../")
        
        form = ExcelImportForm()
        payload = {"form": form}
        return render(request, "admin/excel_form.html", payload)

    def import_cost(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                if not excel_file.name.endswith(('.xlsx', '.xls')):
                    messages.error(request, 'File không phải là Excel (.xlsx)')
                    return HttpResponseRedirect("../")

                try:
                    workbook = load_workbook(excel_file)
                    cost_rows = self._extract_cost_rows(workbook)
                    updated_count, skipped_not_found_count, skipped_missing_code_count = self._import_cost_rows(cost_rows)

                    messages.success(request, f'Cập nhật thành công unit_cost cho {updated_count} item.')
                    if skipped_not_found_count:
                        messages.warning(request, f'Bỏ qua {skipped_not_found_count} item không tìm thấy trong ProductData (item_code không khớp).')
                    if skipped_missing_code_count:
                        messages.warning(request, f'Bỏ qua {skipped_missing_code_count} dòng thiếu Item Code.')
                except Exception as e:
                    messages.error(request, f'Lỗi khi xử lý file Excel: {str(e)}')

                return HttpResponseRedirect("../")

        form = ExcelImportForm()
        payload = {"form": form}
        return render(request, "admin/csv_form.html", payload)

    def _extract_cost_rows(self, workbook):
        """Đọc 2 sheet P2/P3, lọc còn 2 cột (Item Code, Unit Cost),
        trả về danh sách các dòng dữ liệu (đã bỏ dòng header)."""
        target_sheet_names = ['Physical Inventory Result (P2)', 'Physical Inventory Result (P3)']
        kept_sheets = [s for s in workbook.sheetnames if s in target_sheet_names]

        if not kept_sheets:
            raise ValueError('Không tìm thấy sheet cần xử lý')

        p2_rows = []
        p3_rows = []

        for sheet_name in kept_sheets:
            source_sheet = workbook[sheet_name]
            rows = list(source_sheet.iter_rows(values_only=True))
            rows = rows[8:]
            filtered_rows = []
            for row_idx, row in enumerate(rows, start=1):
                if row_idx in {2, 3, 4}:
                    continue
                if row is None:
                    continue
                cleaned_row = list(row[0:9]) if len(row) > 0 else []
                if len(cleaned_row) < 9:
                    cleaned_row.extend([''] * (9 - len(cleaned_row)))
                # Chỉ giữ lại cột 1 và 8 (Item Code, Unit Cost)
                cleaned_row = [cleaned_row[i] for i in range(9) if i not in {0, 2, 3, 4, 5, 6, 7}]
                filtered_rows.append(cleaned_row)

            if sheet_name == 'Physical Inventory Result (P2)':
                p2_rows = filtered_rows
            elif sheet_name == 'Physical Inventory Result (P3)':
                p3_rows = filtered_rows[1:] if filtered_rows else []

        all_rows = p2_rows + p3_rows
        # Dòng đầu tiên (từ P2) là header chữ "Item Code / Unit Cost" -> bỏ qua
        return all_rows[1:] if len(all_rows) > 1 else []

    def _import_cost_rows(self, cost_rows):
        """Cập nhật unit_cost trong bảng ProductData dựa trên item_code. Bỏ qua nếu item_code
        không tồn tại trong ProductData hoặc thiếu Item Code."""
        updated_count = 0
        skipped_not_found_count = 0
        skipped_missing_code_count = 0

        # Gom unit_cost mới nhất theo item_code (dòng sau ghi đè dòng trước nếu trùng)
        cost_by_code = {}
        for row in cost_rows:
            item_code = str(row[0] or '').strip() if len(row) > 0 else ''
            raw_cost = row[1] if len(row) > 1 else None

            if not item_code:
                skipped_missing_code_count += 1
                continue

            try:
                unit_cost = Decimal(str(raw_cost)) if raw_cost not in (None, '') else None
            except (InvalidOperation, ValueError, TypeError):
                unit_cost = None

            cost_by_code[item_code] = unit_cost

        existing_products = list(ProductData.objects.filter(item_code__in=cost_by_code.keys()))
        found_codes = {product.item_code for product in existing_products}
        skipped_not_found_count = len([code for code in cost_by_code if code not in found_codes])

        products_to_update = []
        for product in existing_products:
            product.unit_cost = cost_by_code[product.item_code]
            products_to_update.append(product)

        if products_to_update:
            with transaction.atomic():
                ProductData.objects.bulk_update(products_to_update, ['unit_cost'])
            updated_count = len(products_to_update)

        return updated_count, skipped_not_found_count, skipped_missing_code_count
    def _clean_cell_value(self, cell):
        """Helper method để làm sạch giá trị từ cell Excel"""
        return str(cell.value or '').strip()

admin.site.unregister(User)
admin.site.register(User, CustomUserAdmin)
admin.site.register(Item, ItemAdmin)